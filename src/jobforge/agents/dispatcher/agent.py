"""
JobForge AI — Dispatcher Agent.

Deterministic agent (not Deep): no LLM reasoning needed.
Builds the Excel digest, packages tailored CVs, and sends the email.
"""

from __future__ import annotations

import smtplib
from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any

import pandas as pd
import structlog
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from jobforge.agents.base import BaseAgent
from jobforge.config.settings import OUTPUT_DIR, settings
from jobforge.models.cv import TailoredCV
from jobforge.models.job import ScoredJob
from jobforge.models.scoring import MatchSummary
from jobforge.models.state import JobForgeState

logger = structlog.get_logger(__name__)


class DispatcherAgent(BaseAgent):
    """Packages the daily digest Excel + tailored CVs and sends via email."""

    name = "dispatcher"

    async def run(self, state: JobForgeState) -> dict[str, Any]:
        """Build Excel, send email, return state update."""
        qualified = state.get("qualified_jobs", [])
        tailored_cvs = state.get("tailored_cvs", [])
        match_summary = state.get("match_summary", MatchSummary())

        # Create today's output directory
        today = date.today().isoformat()
        output_dir = OUTPUT_DIR / today
        output_dir.mkdir(parents=True, exist_ok=True)
        (output_dir / "cvs").mkdir(exist_ok=True)

        # Build Excel digest
        excel_path = output_dir / "JobForge_Digest.xlsx"
        self._build_excel(qualified, tailored_cvs, excel_path)

        # Send email
        email_sent = False
        try:
            email_sent = self._send_email(
                excel_path=excel_path,
                cv_paths=[Path(cv.pdf_path) for cv in tailored_cvs if Path(cv.pdf_path).exists()],
                summary=match_summary,
                date_str=today,
            )
        except Exception as e:
            logger.error("dispatcher.email.failed", error=str(e))

        return {
            "excel_path": str(excel_path),
            "email_sent": email_sent,
        }

    def _build_excel(
        self,
        qualified: list[ScoredJob],
        tailored_cvs: list[TailoredCV],
        output_path: Path,
    ) -> None:
        """Build a professionally formatted Excel digest."""
        cv_map = {cv.job_id: cv for cv in tailored_cvs}

        rows = []
        for rank, job in enumerate(qualified, 1):
            cv = cv_map.get(job.job.job_id)
            rows.append({
                "Rank": rank,
                "Match %": round(job.overall_score, 1),
                "Job Title": job.job.title,
                "Company": f"{job.job.company} [{job.job.company_stage}]" if job.job.company_stage and job.job.company_stage != "unknown" else job.job.company,
                "Location": f"{job.job.location}, {job.job.work_model or ''}".strip(", "),
                "Salary": job.job.salary_display,
                "Key Match Reasons": ", ".join(job.key_matching_skills[:3]),
                "Gaps": ", ".join(job.key_gaps[:3]) if job.key_gaps else "None identified",
                "CV Variant": cv.variant_used if cv else "—",
                "Visa": job.visa_tag,
                "Startup": "Yes" if job.job.is_startup else "",
                "Sponsorship": "Yes" if job.job.offers_sponsorship else ("No" if job.job.citizens_only else "Unknown"),
                "Source": job.job.source,
                "Apply URL": job.job.url,
            })

        df = pd.DataFrame(rows)
        df.to_excel(str(output_path), index=False, sheet_name="Job Digest")

        # Apply formatting
        self._format_excel(output_path, len(rows))

        logger.info("dispatcher.excel.created", path=str(output_path), rows=len(rows))

    def _format_excel(self, path: Path, row_count: int) -> None:
        """Apply conditional formatting to the Excel digest."""
        from openpyxl import load_workbook

        wb = load_workbook(str(path))
        ws = wb.active

        # Header formatting
        header_fill = PatternFill("solid", fgColor="1B2A4A")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data formatting
        green_fill = PatternFill("solid", fgColor="E8F5E9")
        amber_fill = PatternFill("solid", fgColor="FFF8E1")

        for row in ws.iter_rows(min_row=2, max_row=row_count + 1):
            score_cell = row[1]  # Match % column
            try:
                score = float(score_cell.value)
                if score >= 85:
                    score_cell.fill = green_fill
                elif score >= 70:
                    score_cell.fill = amber_fill
            except (ValueError, TypeError):
                pass

            for cell in row:
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Column widths
        widths = [6, 8, 30, 22, 18, 16, 30, 25, 12, 14, 8, 12, 12, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(str(path))

    def _send_email(
        self,
        excel_path: Path,
        cv_paths: list[Path],
        summary: MatchSummary,
        date_str: str,
    ) -> bool:
        """Send the daily digest email with attachments."""
        cfg = settings.email
        if not cfg.user or not cfg.password:
            logger.warning("dispatcher.email.no_credentials")
            return False

        # Build subject
        top_company = summary.highest_score_company or "N/A"
        subject = (
            f"JobForge AI | {date_str} | "
            f"{summary.total_qualified} Qualified Jobs | "
            f"Top: {summary.highest_score:.0f}% at {top_company}"
        )

        # Build body
        body = f"""JobForge AI Daily Digest — {date_str}
{'='*50}

Pipeline Summary:
  Total scraped:    {summary.total_scraped}
  After dedup:      {summary.total_after_dedup}
  LLM scored:       {summary.total_scored}
  Qualified (≥70%): {summary.total_qualified}
  Sponsoring roles: {summary.sponsoring_jobs_count}
  Startup roles:    {summary.startup_jobs_count}

Score Distribution:
  90-100%: {summary.score_distribution.get('90-100', 0)}
  80-89%:  {summary.score_distribution.get('80-89', 0)}
  70-79%:  {summary.score_distribution.get('70-79', 0)}

Top Match: {summary.highest_score:.0f}% at {top_company}

Attached: Excel digest + {len(cv_paths)} tailored CVs.

— JobForge AI (Autonomous Pipeline)
"""

        msg = MIMEMultipart()
        msg["From"] = cfg.user
        msg["To"] = cfg.recipient_email or cfg.user
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        # Attach Excel
        if excel_path.exists():
            with open(excel_path, "rb") as f:
                att = MIMEApplication(f.read(), Name=excel_path.name)
                att["Content-Disposition"] = f'attachment; filename="{excel_path.name}"'
                msg.attach(att)

        # Attach CVs (limit to top 10 to avoid email size issues)
        for cv_path in cv_paths[:10]:
            if cv_path.exists():
                with open(cv_path, "rb") as f:
                    att = MIMEApplication(f.read(), Name=cv_path.name)
                    att["Content-Disposition"] = f'attachment; filename="{cv_path.name}"'
                    msg.attach(att)

        # Send
        with smtplib.SMTP(cfg.host, cfg.port) as server:
            server.starttls()
            server.login(cfg.user, cfg.password)
            server.send_message(msg)

        logger.info("dispatcher.email.sent", to=cfg.recipient_email or cfg.user)
        return True
